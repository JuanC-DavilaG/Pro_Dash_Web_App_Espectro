# import io
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import date
from datetime import datetime
import pandas as pd
from scipy.optimize import fsolve
# import pylab
import numpy as np
import matplotlib.pyplot as plt

def Trazar(frecuencias, Anchos_de_banda, Potencias):

    frecuencia_angular = []

    cotas_superiores = []

    cotas_inferiores = []

    fases = []

    frecuencia_angular = [2000*np.pi/BW for BW in Anchos_de_banda]

    inv_frecuencia_angular = np.array([np.pi/invFA for invFA in frecuencia_angular])

    cotas_inferiores = frecuencias - inv_frecuencia_angular

    cotas_superiores = frecuencias + inv_frecuencia_angular

    fases = np.multiply(frecuencias,frecuencia_angular)
    
    
    return frecuencia_angular[0], cotas_inferiores[0], cotas_superiores[0], fases[0]

# Compensamos el error de los valores equiseparados
def equis(a, b, num):
    return np.linspace(a+(b-a)/(2*num), b-(b-a)/(2*num), num)

# Encontramos los puntos de inteseccion de las funciones no lineales
def Buscar_inteseccion(A1, fun1, f1, p1, ofs1, A2, fun2, f2, p2, ofs2, x0):
     return fsolve(lambda x : (A1*fun1(f1*x + p1) + ofs1) - (A2*fun2(f2*x + p2) + ofs2),x0)

# Aproximamos el area de la curva definida en el intervalo [a, b] por medio de rectangulos 
def integral(func, a, b, num):
    return np.abs(np.sum(func)*(a-b)/num)
    
def PorInterferencia(frecuencias, Potencias, frecuencia_angular, cotas_inferiores, cotas_superiores, fases, posI, posE):

    fec0=frecuencias[posI]
    Amp0=Potencias[posI]/2
    frec0=frecuencia_angular[posI]
    pha0=fases[posI]
    ofst0=Potencias[posI]/2
    cotI0=cotas_inferiores[posI]
    cotS0=cotas_superiores[posI]

    fec1=frecuencias[posE]
    Amp1=Potencias[posE]/2
    frec1=frecuencia_angular[posE]
    pha1=fases[posE]
    ofst1=Potencias[posE]/2
    cotI1=cotas_inferiores[posE]
    cotS1=cotas_superiores[posE]

    num = 100

    # Punto de inteseccion entre las curvas
    inicial = fec0-np.abs(cotI0-fec0)/2 if posI == 0 else fec1+np.abs(fec1-cotS1)/2

    Pcorte = Buscar_inteseccion(Amp0, np.cos, frec0, -pha0, ofst0, Amp1, np.cos, frec1, -pha1, ofst1, inicial)

    # Curva de interferencia lado invasor
    x = equis(Pcorte, cotI0, num) if posI == 0 else equis(Pcorte, cotS1, num)

    fx = (Amp0)*(1 + np.cos(frec0*x - pha0)) if posI == 0 else (Amp1)*(1 + np.cos(frec1*x - pha1))


    # Curva de interferencia lado invadido
    t = equis(cotS1, Pcorte, num) if posI == 0 else equis(cotI0, Pcorte, num)

    gt = (Amp1)*(1 + np.cos(frec1*t - pha1)) if posI == 0 else (Amp0)*(1 + np.cos(frec0*t - pha0))


    # Areas interferidas (En solapamiento)
    areaFxinterfd = integral(fx, Pcorte, cotS1, num) if posI == 0 else integral(fx, Pcorte, cotS1, num)

    areaGtinterft = integral(gt, cotI0, Pcorte, num) if posI == 0 else integral(gt, cotI0, Pcorte, num)


    EsIn = float(areaFxinterfd) + float(areaGtinterft)

    # Portadora interferida
    z = equis(cotS0, cotI0, num) if posI == 0 else equis(cotS1, cotI1, num)

    areaFx = integral((Amp0)*(1 + np.cos(frec0*z - pha0)), cotI0, cotS0, num) if posI == 0 else integral((Amp1)*(1 + np.cos(frec1*z - pha1)), cotI1, cotS1, num)

    return EsIn*100/areaFx
    

def AnalisDePortadoras(TipoFrec, datf, datg):

    listIntererencias = []
    numDes = 5

    # Primer ciclo para recorres todas las fecuencias propuestas
    for j in range(len(datg[TipoFrec])):
        
        #Obtenemos los valores de frecuencia y potencia de la primera frecuencia propuesta
        valor = np.round(datg[TipoFrec].loc[j], decimals = numDes)
        Pwl = datg[" P.I.R.E. (dBW)"].loc[j]
        
        #Encontramos el valor mas cercano
        PosNum = (datf["Frecuencias"] - valor).abs().idxmin()
        numero = np.round(datf["Frecuencias"].loc[PosNum], decimals=4)

        AnB = datf["Anchos de banda"].loc[PosNum]

        #Obtenemos sus limites laterales
        limB = np.round((numero - AnB/2000), decimals=numDes)
        limA = np.round((numero + AnB/2000), decimals=numDes)

        #Obtenemos los valores para trazar la curva
        FanguB_V, limB_V, limA_V, phaB_V = Trazar(
                                            [valor], 
                                            [datg["Ancho de Banda"].loc[j]], 
                                            [Pwl])
        
        indis = set()
        #Recorremos todas las frecuencias conceciondas desde la mas sercana hacia abajo
        for i in range(PosNum, 0, -1):
            
            fr = np.round(datf["Frecuencias"].loc[i], decimals = numDes)

            AnB = datf["Anchos de banda"].loc[i]

            PwO = datf["P.I.R.E (dBW)"].loc[i]

            FanguB, limB, limA, phaB = Trazar([fr], [AnB], [PwO])

            if valor == fr: 

                indis.add(i)

            #Evaluamos si la propuesta esta por delante de la concecion
            elif valor - fr > 0:

                #Evaluamos si el limite superior de la concecion esta dentro de
                #los limetes de la propuesta.
                if limB_V < limA and limA < limA_V:
                    
                    listIntererencias.append(valor)
                    
                    tupInterferido = [i, 
                                        datf["#"].loc[i], 
                                        fr, AnB, PwO, 
                                        datf["razon social"].loc[i], 
                                        datg["#"].loc[j]]

                    
                    porsen = PorInterferencia([fr, valor], 
                                            [PwO, Pwl], 
                                            [FanguB, FanguB_V], 
                                            [limB, limB_V], 
                                            [limA, limA_V], 
                                            [phaB, phaB_V], 1, 0)
                    
                    tupInterferido.append(str(np.round(porsen, decimals=3)) + " %")
                    
                    listIntererencias.append(tuple(tupInterferido))

                    indis.add(i)

                else:
                    break

        #Recorremos todas la frecuencias conceciondas desde la mas secana hacia arriba
        for i in range(PosNum, len(datf)):
            
            # indi = PosNum + k

            fr = np.round(datf["Frecuencias"].loc[i], decimals = numDes)

            AnB = datf["Anchos de banda"].loc[i]

            PwO = datf["P.I.R.E (dBW)"].loc[i]

            FanguB, limB, limA, phaB = Trazar([fr], [AnB], [1])

            if valor == fr: 

                indis.add(i)

            #Evaluamos si la propuesta esta por detras de la concecion
            elif valor - fr < 0:

                #Evaluamos si el limite inferior de la concecion esta dentro de
                #los limetes de la propuesta.
                if limA_V > limB and limA_V < limA:
                    
                    listIntererencias.append(valor)

                    tupInterferido = [i, 
                                        datf["#"].loc[i], 
                                        fr, AnB, PwO, 
                                        datf["razon social"].loc[i],
                                        datg["#"].loc[j]]

                    porsen = PorInterferencia([valor, fr], 
                                            [Pwl, PwO], 
                                            [FanguB_V, FanguB], 
                                            [limB_V, limB], 
                                            [limA_V, limA], 
                                            [phaB_V, phaB], 1, 0)
                    
                    tupInterferido.append(str(np.round(porsen, decimals=3)) + " %")
                    
                    listIntererencias.append(tuple(tupInterferido))
                    
                    indis.add(i)

                else:
                    break

    return listIntererencias


def EstudioDeInvacion(datf, datg):

        interferenciaTx = AnalisDePortadoras("Frecuencias Tx", datf, datg)
        interferenciaRx = AnalisDePortadoras("Frecuencias Rx", datf, datg)

        interferencia = interferenciaTx + interferenciaRx

        return RepoAnalisis( interferencia )


def RepoAnalisis( interferencia ):

    timestamp = datetime.now()
    format = timestamp

    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"

    ws['A3'].value = 'Propuestas'
    ws['A3'].alignment = Alignment(horizontal='center')

    ws['A4'].value = '#'
    ws['B4'].value = 'Frecuencia'

    ws['C3'].value = 'Codición espectral'
    ws['C3'].alignment = Alignment(horizontal='center')

    ws['C4'].value = 'Espectro Libre'
    ws['D4'].value = 'Espectro Protegido'
    ws['E4'].value = 'Espectro Móvil Maritimo'
    ws['F4'].value = 'CNAF'
    ws['G4'].value = 'Otros'

    ws['H3'].value = 'Interferencia lateral'
    ws['H3'].alignment = Alignment(horizontal='center')

    ws['H4'].value = '# de coincidencias'
    ws['I4'].value = "# de Id"
    ws['J4'].value = "Frecuencias"
    ws['K4'].value = 'Concesionarios / Autorizados'

    ws['L4'].value = '% de invación'

    j=0
    for i in range(len(interferencia)):

        if(i%2 == 0 and i < len(interferencia)-1):

            ws['A'+str(5+i-j)].value = str(interferencia[i+1][6])
            ws['B'+str(5+i-j)].value = str(interferencia[i])
            ws['I'+str(5+i-j)].value = str(interferencia[i+1][1])
            ws['J'+str(5+i-j)].value = str(interferencia[i+1][2])
            ws['K'+str(5+i-j)].value = str(interferencia[i+1][5])
            ws['L'+str(5+i-j)].value = str(interferencia[i+1][7])

            j=j+1

    return pd.DataFrame(ws.values, columns = [format,'','','','','','','','','','',''])