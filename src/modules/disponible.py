import pandas as pd
import numpy as np
from openpyxl import Workbook
from modules.dbConnect import *

# Motor Espectro
def Ocupacion(band, can, baja, alta, condiciones, SerPer, concecionado):
    
    # Leemos los datos
    csv_tablas = traerData()

    ################ Inicio construir segmento

    arr = []

    com = 1000

    n = int(((alta-baja)/can)*com)+1

    j = baja*com

    for i in range(0 , n):

        arr.append(j/com)

        j = j + can


    arr = np.array(arr)

    ################ Fin construir segmento


    ################ Inicio Espectro Libre

    if('L' in condiciones):

        ################ Inicio Buscar Espectro Libre

        Lb_B = csv_tablas[csv_tablas[band+'.Lib.B'].notna()][band+'.Lib.B']

        Lb_A =  csv_tablas[csv_tablas[band+'.Lib.A'].notna()][band+'.Lib.A']

        Libre = list(zip(Lb_B, Lb_A))

        ################ Fin Buscar Espectro Libre

        ################ Inicio Quitar Espectro Libre

        ELibre = []

        for i in range(0, len(Libre)):

            ELibre.append(np.where((arr >= Libre[i][0]) & (arr <= Libre[i][1])))

            arr = np.delete(arr, ELibre[i], axis=None)

    else:
        Libre = ['N/A']

        ################ Fin Quitar Espectro Libre

    ################ Fin Espectro Libre


    ################ Inicio Espectro Protegido

    if('P' in condiciones):

        try:
        ################ Inicio Buscar Espectro Protegido

            Protegido = list(csv_tablas[csv_tablas[band+'.Pro'].notna()][band+'.Pro'])

        ################ Fin Buscar Espectro Protegido


        ################ Inicio Quitar Espectro Protegido

            # Busca cada valor en el arreglo del espectro canalizado y descarta los valores en
            # espectro protegito, arroja una tupla con los indices de los valos a descartar y su
            # tipo por lo que es necesario extraer el arreglo con el primer [0] y posteriormente 
            # el valor con el segundo [0].
            for valorPotegido in Protegido:
                arr = np.delete(arr, np.where(arr == valorPotegido)[0][0], axis=0)

        ################ Fin Quitar Espectro Protegido

        except:
            Protegido = ['N/A']
            pass

    else:
        Protegido = ['N/A']

    ################ Fin Espectro Protegido


    ################ Inicio Servicio Móvil Maritimo

    if('S' in condiciones):

        try:

            ################ Inicio Buscar Espectro Servicio Móvil Maritimo

            Smm_B = csv_tablas[csv_tablas[band+'.SMM.B'].notna()][band+'.SMM.B']

            Smm_A =  csv_tablas[csv_tablas[band+'.SMM.A'].notna()][band+'.SMM.A']

            Maritimo = list(zip(Smm_B, Smm_A))

        ################ Fin Buscar Espectro Servicio Móvil Maritimo

        ################ Inicio Quitar Espectro Servicio Móvil Maritimo

            ESMM = []

            for i in range(0, len(Maritimo)):

                ESMM.append(np.where((arr >= Maritimo[i][0]) & (arr <= Maritimo[i][1])))

                arr = np.delete(arr, ESMM[i], axis=None)

        ################ Fin Quitar Espectro Servicio Móvil Maritimo

        except:
            Maritimo = ['N/A']
            pass

    else:
        Maritimo = ['N/A']


    ################ Fin Espectro Servicio Móvil Maritimo

    ################ Inicio Espectro de Frontera *********

    if('F' in condiciones):

        try:
        ################ Inicio Buscar Espectro de Frontera

            Frontera = list(csv_tablas[csv_tablas[band+'.Fron'].notna()][band+'.Fron'])

        ################ Fin Buscar Espectro de Frontera


        ################ Inicio Quitar Espectro de Frontera

            # Busca cada valor en el arreglo del espectro canalizado y descarta los valores en
            # espectro en frontera, en este caso no es necesario extraer de un arreglo ya que en
            # eventos posteriores se habia realizado.
            for valorFrontera in Frontera:
                arr = np.delete(arr, np.where(arr == valorFrontera))

        ################ Fin Quitar Espectro de Frontera

        except:
            Frontera = ['N/A']
            pass
    else:
        Frontera = ['N/A']

    ################ Fin Espectro Frontera *********


    ################ Inicio CNAF

    if('C' in condiciones):

        ################ Inicio Buscar CNAF

        # Buscamos los limites el los que se cumple el segmento.
        Lmin  = np.where(csv_tablas[band+'.Cnaf.B'] >= baja)[0][0]

        Lmax  = len(np.where(csv_tablas[band+'.Cnaf.A'] <= alta)[0])

        # Columna de los servicios en el CNAF.
        lisevis = list(csv_tablas[band+'.Cnaf.Ser'][Lmin:Lmax])

        # Buscamos si es que existe la palabra Móvil en la columna '>' o buscamos las que no pertenecen '=='.
        Noservis=[]

        for i in range(0, len(SerPer)):
            SerTemp = [j+Lmin for j in range(0, len(lisevis)) if lisevis[j].split(" | ").count(SerPer[i]) > 0]

            for j in range(0, len(SerTemp)): Noservis.append(SerTemp[j])

        # Eliminamos las columnas que estan dadas de alta al servicio movil y reiniciamos los indices.
        csv_tablas = pd.DataFrame(csv_tablas).drop(Noservis,axis=0).reset_index(drop=True)

        # Buscamos los nuevos limites que cumplan con el segmento.
        Lmin  = np.where(csv_tablas[band+'.Cnaf.B'] >= baja)[0][0]

        Lmax  = len(np.where(csv_tablas[band+'.Cnaf.A'] <= alta)[0])

        # Buscamos los limites el las columanas de los segmentos altos y bajos y creamos sublistas con estos elementos.
        segBajas = csv_tablas[band+'.Cnaf.B'][Lmin:Lmax]

        segAltas = csv_tablas[band+'.Cnaf.A'][Lmin:Lmax]

        # Creamos la estructura de los datos como una lista de tuplas para los segmentos.
        CNAF = list(zip(segBajas,segAltas))

        # Creamos la estructura de los datos como una lista de tuplas para los servicios.
        movil = [
            tuple(list(csv_tablas[band+'.Cnaf.Ser'][Lmin:Lmax])[i].split(" | ")) 
            for i in range(0, len(csv_tablas[band+'.Cnaf.Ser'][Lmin:Lmax]))
        ]

        ################ Fin Buscar CNAF

        ################ Inicio Quitar CNAF

        Ecnaf = []

        for i in range(0, len(CNAF)):

            Ecnaf.append(np.where((arr >= CNAF[i][0]) & (arr <= CNAF[i][1])))

            arr = np.delete(arr, Ecnaf[i], axis=None)

        if(len(CNAF) == 0): CNAF = ["Sin segmentos no autorizados."]
        if(len(movil) == 0): movil = ["Sin servicios no autorizados. "]
    
    else:
        CNAF = ['N/A']
        movil = ['N/A']

        ################ Fin Quitar CNAF

    ################ Fin CNAF

    ################ Inicio Espectro Concecionado

    try:

    ################ Inicio Quitar Espectro Concecionado
        for i in range(0, len(concecionado['Frecuencias'])):
            try:
                arr = np.delete(arr, np.where(arr == concecionado['Frecuencias'][i])[0][0], axis=0)
            except:
                pass

    ################ Fin Quitar Espectro concecionado

    except:
        Concecionado = ['N/A']
        pass

    ################ Fin Espectro Concecionado

    return InfoOcupacion(arr, Libre, Protegido, Maritimo, Frontera, CNAF, movil, concecionado)

def InfoOcupacion(arr, Libre, Protegido, Maritimo, Frontera, CNAF, Servicios, Concecionado):

    wb = Workbook()
    ws = wb.active

    ofs = 1
    
    for row in range(0, len(arr)):
        ws.cell(row+ofs, 1).value = arr[row]
        
    for row in range(0, len(Concecionado['Frecuencias'])):
        ws.cell(row+ofs, 2).value = Concecionado['Frecuencias'][row]

    for row in range(0, len(Libre)):
        ws.cell(row+ofs, 3).value = str(Libre[row])

    for row in range(0, len(Protegido)):
        ws.cell(row+ofs, 4).value = Protegido[row]

    for row in range(0, len(Maritimo)):
        ws.cell(row+ofs, 5).value = str(Maritimo[row])
            
    for row in range(0, len(Frontera)):
        ws.cell(row+ofs, 6).value = Frontera[row]

    for row in range(0, len(CNAF)):
        ws.cell(row+ofs, 7).value = str(CNAF[row])
            
        ws.cell(row+ofs, 8).value = str(Servicios[row])

    return pd.DataFrame(ws.values, 
                        columns = ['Disponibilidad',
                                    'Concesionado',
                                    'Espectro Libre',
                                    'Espectro Protegido',
                                    'Sistema Móvil Marítimo',
                                    'Espectro de Frontera',
                                    'CNAF',
                                    'Servicios',
                                ]
                    )