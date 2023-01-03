# import dash
# from dash import html, dcc, html, Input, Output, callback

# dash.register_page(__name__, path='/resultados')

# layout = html.Div(children=[
#     html.H1(children='Esta es nuestra pagina de resultados', id='titulo_resultados'),

#     html.Div(children='''
#         Este es el contenido de la pagina de resultados.
#     ''', id='contenido_resultados'),

# ])

import dash
from dash import html, dcc, html, Input, Output, callback
from dash.dependencies import State
import dash_bootstrap_components as dbc
import plotly.graph_objects as go
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import csv
import matplotlib.pyplot as plt

dash.register_page(__name__, path='/resultados')

layout = html.Div(
    [
        html.Div(id="graf-0", style={"verticalAlign": "middle", "color": "white"}),
        dbc.Button(
            "Leer", id="sub-leer", className="me-2", n_clicks=0
        ),
        html.Div(id='conten', children=[
            html.Br(),
            html.Label(id='por', children='por_1')
        ])
    ]
)

def leer(valor):

    if valor != 0:
        # wb = load_workbook("D:/Users/juan.davila/Documents/Proyectos py/Graficas del espectro/app/2_Reasignador.xlsm")
        # ws_A = wb['Análisis']
        # with open('D:/Users/juan.davila/Pictures/Saved Pictures/Test_Propuesta.csv', newline='') as csvfile:
        #     spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
        # df = pd.read_excel(
        #         open('D:/Users/juan.davila/Pictures/Saved Pictures/reporte_test.xlsx', 'rb'),
        #         sheet_name='Hoja1'
        #     ) 
        wb = load_workbook("D:/Users/juan.davila/Pictures/Saved Pictures/reporte_test.xlsx")
        ws_A = wb["Hoja1"]
        # df = pd.read_excel(
        #      open('D:/Users/juan.davila/Documents/Proyectos py/Graficas del espectro/app/2_Reasignador.xlsm', 'rb'),
        #      sheet_name='Análisis'
        # ) 
            # for row in spamreader:
            #     print(', '.join(row))

    return (ws_A)


@callback(
    Output("graf-0", "children"), 
    [Input("sub-leer", "n_clicks")]
)
def on_button_click(n):
    if n is None:
        return "Not clicked."
    elif (n > 0):

        print(leer(n))
        return leer(n)
    else:
        return f"Clicked {n} times."